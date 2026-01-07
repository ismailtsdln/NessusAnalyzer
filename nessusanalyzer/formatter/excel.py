from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ..core.models import NessusReport
from ..utils.logger import logger

class ExcelFormatter:
    def __init__(self, report: NessusReport):
        self.report = report

    def export(self, output_path: str):
        """Exports the report to an Excel (XLSX) file."""
        logger.info(f"Exporting report to Excel: {output_path}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Nessus Findings"

        # Headers
        headers = [
            "Host", "IP", "Plugin ID", "Plugin Name", "Severity", 
            "Risk Factor", "CVSS Base Score", "Exploit Available", "Metasploit", "CVE"
        ]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data
        severity_colors = {
            4: "FF0000",  # Critical - Red
            3: "FF6600",  # High - Orange
            2: "FFFF00",  # Medium - Yellow
            1: "00B050",  # Low - Green
            0: "C0C0C0"   # Info - Gray
        }

        for host in self.report.hosts:
            for finding in host.findings:
                row_data = [
                    host.name,
                    host.ip or "",
                    finding.plugin_id,
                    finding.plugin_name,
                    finding.severity,
                    finding.risk_factor,
                    finding.cvss_base_score or "",
                    "Yes" if finding.exploit_available else "No",
                    finding.metasploit_name or "",
                    ", ".join(finding.cve)
                ]
                ws.append(row_data)
                
                # Color code severity cell
                severity_cell = ws.cell(row=ws.max_row, column=5)
                color = severity_colors.get(finding.severity, "FFFFFF")
                severity_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)

        try:
            wb.save(output_path)
            logger.info("Excel export completed successfully.")
        except Exception as e:
            logger.error(f"Error during Excel export: {e}")
            raise
